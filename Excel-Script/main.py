# Author: Asim Sheikh (Software Enigneer @ Genwealth 360 Inc.)

# import pandas as pd
# from openpyxl import load_workbook

# # To Load in Excel Sheet
# # file_path = "template.xlsx"
# # file_path = "20241018 Master Template V 8.6.xlsx"
# # workbook = load_workbook(file_path)
# # sheet = workbook.active

# def calculate_se_tax(schedule_c_income, w2_income, year, partnership_income, filing_status):
#     SS_WAGE_BASE = {2022: 147000, 2023: 160200, 2024: 168600}[year]
#     SOCIAL_SECURITY_RATE = 0.124  # 12.4% total 
#     MEDICARE_RATE = 0.029         # 2.9% total
#     SE_INCOME_FACTOR = 0.9235     # Adjustment factor for self-employment income

#     # Adjust incomes
#     adjusted_schedule_c_income = schedule_c_income * SE_INCOME_FACTOR
#     adjusted_partnership_income = partnership_income * SE_INCOME_FACTOR
#     total_adjusted_se_income = adjusted_schedule_c_income + adjusted_partnership_income

#     # Calculate Social Security Tax (for both W2 and self-employment)
#     total_ss_wages = min(w2_income, SS_WAGE_BASE)
#     ss_taxable_limit = SS_WAGE_BASE - total_ss_wages
#     ss_taxable_se_income = min(total_adjusted_se_income, ss_taxable_limit)
    
#     # Social Security tax for W2 and self-employment income
#     social_security_tax_w2 = total_ss_wages * SOCIAL_SECURITY_RATE
#     social_security_tax_se = ss_taxable_se_income * SOCIAL_SECURITY_RATE

#     # Calculate Medicare Tax (for both W2 and self-employment - no wage base limit)
#     medicare_tax_w2 = w2_income * MEDICARE_RATE
#     medicare_tax_se = total_adjusted_se_income * MEDICARE_RATE

#     # Total wages including W2 and Self-Employment for Additional Medicare Tax
#     total_medicare_wages = w2_income + total_adjusted_se_income

#     # Additional Medicare Tax based on Filing Status
#     additional_medicare_income = 0
#     if filing_status == 'single' and total_medicare_wages > 200000:
#         additional_medicare_income = total_medicare_wages - 200000
#     elif filing_status == 'married_jointly' and total_medicare_wages > 250000:
#         additional_medicare_income = total_medicare_wages - 250000
#     elif filing_status == 'married_separately' and total_medicare_wages > 125000:
#         additional_medicare_income = total_medicare_wages - 125000
#     elif filing_status == 'head_of_household' and total_medicare_wages > 200000:
#         additional_medicare_income = total_medicare_wages - 200000
#     elif filing_status == 'qualifying_widow' and total_medicare_wages > 200000:
#         additional_medicare_income = total_medicare_wages - 200000

#     # Additional Medicare Tax
#     additional_medicare_tax = additional_medicare_income * 0.009

#     # Total Self-Employment Tax (including W2 and self-employment, plus the additional Medicare tax)
#     total_se_tax = social_security_tax_w2 + social_security_tax_se + medicare_tax_w2 + medicare_tax_se + additional_medicare_tax

#     return {
#         'adjusted_schedule_c_income': adjusted_schedule_c_income,
#         'adjusted_partnership_income': adjusted_partnership_income,
#         'social_security_tax_w2': social_security_tax_w2,
#         'social_security_tax_se': social_security_tax_se,
#         'medicare_tax_w2': medicare_tax_w2,
#         'medicare_tax_se': medicare_tax_se,
#         'additional_medicare_tax': additional_medicare_tax,
#         'total_se_tax': total_se_tax,
#         'total_adjusted_se_income': total_adjusted_se_income
#     }

# # Test Case:
# client_partnership_income = 500000
# client_schedule_c_income = 3000000
# client_w2_income = 750000
# year = 2024
# filing_status = 'married_separately'

# se_tax_details = calculate_se_tax(
#     client_schedule_c_income,
#     client_w2_income,
#     year,
#     partnership_income=client_partnership_income,
#     filing_status=filing_status
# )

# print(f"---------------------------------------") 
# print(f"Adjusted Schedule C Income: ${se_tax_details['adjusted_schedule_c_income']:,.2f}")
# print(f"Adjusted Partnership Income: ${se_tax_details['adjusted_partnership_income']:,.2f}")
# print(f"Social Security Tax (W2): ${se_tax_details['social_security_tax_w2']:,.2f}")
# print(f"Social Security Tax (Self-Employment): ${se_tax_details['social_security_tax_se']:,.2f}")
# print(f"Medicare Tax (W2): ${se_tax_details['medicare_tax_w2']:,.2f}")
# print(f"Medicare Tax (Self-Employment): ${se_tax_details['medicare_tax_se']:,.2f}")
# print(f"Additional Medicare Tax: ${se_tax_details['additional_medicare_tax']:,.2f}")
# print(f"Total Self-Employment Tax: ${se_tax_details['total_se_tax']:,.2f}")
# print(f"---------------------------------------")

import openpyxl

# Function to calculate self-employment tax
def calculate_se_tax(schedule_c_income, w2_income, year, partnership_income, filing_status):
    SS_WAGE_BASE = {2022: 147000, 2023: 160200, 2024: 168600}[year]
    SOCIAL_SECURITY_RATE = 0.124  # 12.4% total 
    MEDICARE_RATE = 0.029         # 2.9% total
    SE_INCOME_FACTOR = 0.9235     # Adjustment factor for self-employment income

    # Adjust incomes
    adjusted_schedule_c_income = schedule_c_income * SE_INCOME_FACTOR
    adjusted_partnership_income = partnership_income * SE_INCOME_FACTOR
    total_adjusted_se_income = adjusted_schedule_c_income + adjusted_partnership_income

    # Calculate Social Security Tax (for both W2 and self-employment)
    total_ss_wages = min(w2_income, SS_WAGE_BASE)
    ss_taxable_limit = SS_WAGE_BASE - total_ss_wages
    ss_taxable_se_income = min(total_adjusted_se_income, ss_taxable_limit)
    
    # Social Security tax for W2 and self-employment income
    social_security_tax_w2 = total_ss_wages * SOCIAL_SECURITY_RATE
    social_security_tax_se = ss_taxable_se_income * SOCIAL_SECURITY_RATE

    # Calculate Medicare Tax (for both W2 and self-employment - no wage base limit)
    medicare_tax_w2 = w2_income * MEDICARE_RATE
    medicare_tax_se = total_adjusted_se_income * MEDICARE_RATE

    # Total wages including W2 and Self-Employment for Additional Medicare Tax
    total_medicare_wages = w2_income + total_adjusted_se_income

    # Additional Medicare Tax based on Filing Status
    additional_medicare_income = 0
    if filing_status == 'single' and total_medicare_wages > 200000:
        additional_medicare_income = total_medicare_wages - 200000
    elif filing_status == 'married_jointly' and total_medicare_wages > 250000:
        additional_medicare_income = total_medicare_wages - 250000
    elif filing_status == 'married_separately' and total_medicare_wages > 125000:
        additional_medicare_income = total_medicare_wages - 125000
    elif filing_status == 'head_of_household' and total_medicare_wages > 200000:
        additional_medicare_income = total_medicare_wages - 200000
    elif filing_status == 'qualifying_widow' and total_medicare_wages > 200000:
        additional_medicare_income = total_medicare_wages - 200000

    # Additional Medicare Tax
    additional_medicare_tax = additional_medicare_income * 0.009

    # Total Self-Employment Tax (including W2 and self-employment, plus the additional Medicare tax)
    total_se_tax = social_security_tax_w2 + social_security_tax_se + medicare_tax_w2 + medicare_tax_se + additional_medicare_tax

    return {
        'adjusted_schedule_c_income': adjusted_schedule_c_income,
        'adjusted_partnership_income': adjusted_partnership_income,
        'social_security_tax_w2': social_security_tax_w2,
        'social_security_tax_se': social_security_tax_se,
        'medicare_tax_w2': medicare_tax_w2,
        'medicare_tax_se': medicare_tax_se,
        'additional_medicare_tax': additional_medicare_tax,
        'total_se_tax': total_se_tax,
        'total_adjusted_se_income': total_adjusted_se_income
    }

# Function to process Excel spreadsheet
def process_client_spreadsheet(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Assuming data is in the first sheet

    # Iterate through rows (starting from row 2, skipping the header)
    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):  # start=2 to give the correct row number
        client_id = row[0].value
        client_name = row[1].value
        year = int(row[2].value) if row[2].value else 2024
        schedule_c_income = row[3].value
        w2_income = row[4].value
        partnership_income = row[5].value
        filing_status = row[6].value.lower() if row[6].value else "single"

        # Skip rows where income or necessary info is missing
        if None in [schedule_c_income, w2_income, partnership_income]:
            continue

        # Calculate the self-employment tax using the calculate_se_tax function
        se_tax_details = calculate_se_tax(
            schedule_c_income,
            w2_income,
            year,
            partnership_income,
            filing_status
        )

        # Write the calculated tax details back to the spreadsheet
        sheet.cell(row=idx, column=8).value = se_tax_details['adjusted_schedule_c_income']
        sheet.cell(row=idx, column=9).value = se_tax_details['adjusted_partnership_income']
        sheet.cell(row=idx, column=10).value = se_tax_details['social_security_tax_w2']
        sheet.cell(row=idx, column=11).value = se_tax_details['social_security_tax_se']
        sheet.cell(row=idx, column=12).value = se_tax_details['medicare_tax_w2']
        sheet.cell(row=idx, column=13).value = se_tax_details['medicare_tax_se']
        sheet.cell(row=idx, column=14).value = se_tax_details['additional_medicare_tax']
        sheet.cell(row=idx, column=15).value = se_tax_details['total_se_tax']

    # Save the updated workbook
    workbook.save(file_path)

# Example usage
process_client_spreadsheet('Excel-Script\client_data.xlsx')
