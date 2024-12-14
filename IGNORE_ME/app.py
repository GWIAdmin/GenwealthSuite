from flask import Flask, render_template, request
import openpyxl
import os

app = Flask(__name__)

EXCEL_PATH = 'tax_calculations.xlsx'

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Extract data from form
        input_data = {
            'wages': float(request.form.get('wages', 0)),
            'interest': float(request.form.get('interest', 0)),
            # Add more fields as needed
        }

        # Process data with Excel
        results = process_with_excel(input_data)

        # Render results page
        return render_template('results.html', results=results)
    return render_template('index.html')

def process_with_excel(input_data):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Input data into specific cells
    ws['A1'] = input_data['wages']
    ws['A2'] = input_data['interest']
    # Map input_data to appropriate cells

    # Save changes to the workbook
    wb.save(EXCEL_PATH)

    # Retrieve calculated results
    results = {
        'total_income': ws['B10'].value,
        'tax_due': ws['B20'].value,
        # Extract other results as needed
    }

    return results

if __name__ == '__main__':
    app.run(debug=True)